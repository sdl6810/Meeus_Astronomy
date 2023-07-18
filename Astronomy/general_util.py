import os, sys
from openpyxl import load_workbook

#code should determine when it should stop cycling through the column
def importXlColumnToPythonList(filepath, worksheetName, columnLetter, character):
	wbk = load_workbook(filepath)
	sht = wbk[worksheetName]

	data = []
	for i in range(0,(sht.max_row-(sht.min_row+1))):
		if sht[columnLetter + str((sht.min_row+2)+i)].value is None:
			data.append(character)
		else:
			data.append(sht[columnLetter+str((sht.min_row+2)+i)].value)
	return data

# def importXlTable(filepath, worksheetName):
# 	#function should test where the end of the column is


def main():
	storedData = importXlColumnToPythonList('/home/sdl5384/Desktop/Python_SRC/Astronomy/lunar_phases.xlsx','periApo', 'B', 0)

main()