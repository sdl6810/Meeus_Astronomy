import math
from astro_functions import julianDay, reduceAngle, toGregorian
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import matplotlib

def venusLitFraction(julianDay):
	T = round((julianDay - 2451545)/36525,9)
	V = round(reduceAngle(261.51 + 22518.443*T),3)
	M = round(reduceAngle(177.53 + 35999.050*T),3)
	N = round(reduceAngle(50.42 + 58517.811*T),3)
	W = round(V + 1.91*math.sin(math.radians(M)) + 0.78*math.sin(math.radians(N)),6)
	delta = round(math.sqrt(1.52321 + 1.44666*math.cos(math.radians(W))),6)

	k = (math.pow(0.72333 + delta,2)-1)/(2.89332*delta)
	k = round(k,8)

	return [k, delta]

def main():
	wb = Workbook()
	ws = wb.active
	ws['A1'] = "Month"
	ws['B1'] = "Date"
	ws['C1'] = "Year"
	ws['D1'] = "Fraction"
	ws['E1'] = "Julian Day"
	ws['F1'] = 'Earth-Venus Distance'

	phaseAngleDates = []
	phaseAngles = []
	planetaryDistances = []
	for i in range(0,2001):
		jd = julianDay(8,13,2023)-583.921361*i
		phaseAngle = venusLitFraction(jd)[0]
		evDistance = venusLitFraction(jd)[1]
		phaseAngles.append(phaseAngle)
		phaseAngleDates.append(jd)
		planetaryDistances.append(evDistance)
		
		for j in range(0,len(toGregorian(phaseAngleDates[i]))):
			ws[get_column_letter(1+j)+str(2+i)] = toGregorian(phaseAngleDates[i])[j]

		ws[get_column_letter(4)+str(2+i)] = phaseAngles[i]
		ws[get_column_letter(5)+str(2+i)] = phaseAngleDates[i]
		ws[get_column_letter(6)+str(2+i)] = planetaryDistances[i]

	wb.save('/home/sdl5384/Desktop/Inferior Conjunction Dates and Phases (in reverse).xlsx')

main()