import math
from astro_functions import reduceAngle, julianDay
from openpyxl import Workbook, load_workbook
from general_util import importXlColumnToPythonList

def determine_tau(julianDate):
	return round((julianDate - 2451545.0)/365250,12)

def venus_heliocentric_coordinates(jd):
	dataWbk = load_workbook('/home/sdl5384/Desktop/Python_SRC/Astronomy/planetaryData.xlsx')
	dataSht = dataWbk.worksheets[1]
	tau = determine_tau(jd)

	lTerms = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0] 
	bTerms = [0.0, 0.0, 0.0, 0.0]

	print('first loop')
	for i in range(0,24):
		#B and C terms are already in radians so it is not necessary to convert the entire argument to radians
		trigArgument = round(dataSht['D'+str(2+i)].value + dataSht['E'+str(2+i)].value*tau,7)
		lTerms[0] = round(lTerms[0] + dataSht['C'+str(2+i)].value*math.cos(trigArgument),0)
		print(f"{dataSht['C'+str(2+i)].value}*cos({dataSht['D'+str(2+i)].value}+{dataSht['E'+str(2+i)].value}*tau)")

	print('second loop')
	for j in range(0,12):
		#B and C terms already in radians.  No need to convert to radians
		trigArgument = dataSht['D'+str(26+j)].value + dataSht['E'+str(26+i)].value*tau
		lTerms[1] = round(lTerms[1] + dataSht['C'+str(26+j)].value*math.cos(trigArgument),0)
		print(f"{dataSht['C'+str(26+j)].value}*cos({dataSht['D'+str(26+j)].value}+{dataSht['E'+str(26+j)].value}*tau)")

	print('third loop')
	for k in range(0,8):
		#B and C terms are already in radians.
		trigArgument = round(dataSht['D'+str(38+k)].value + dataSht['E'+str(38+k)].value*tau,7)
		lTerms[2] = round(lTerms[2] + dataSht['C'+str(38+k)].value*math.cos(trigArgument),0)
		print(f"{dataSht['C'+str(38+k)].value}*cos({dataSht['D'+str(38+k)].value}+{dataSht['E'+str(38+k)].value}*tau)")

	print('fourth loop')
	for m in range(0,3):
		#B and C terms are already in radians
		trigArgument = round(dataSht['D'+str(46+m)].value + dataSht['E'+str(46+m)].value*tau,7)
		lTerms[3] = round(lTerms[3] + dataSht['C'+str(46+m)].value*math.cos(trigArgument),0)
		print(f"{dataSht['C'+str(46+m)].value}*cos({dataSht['D'+str(46+m)].value}+{dataSht['E'+str(46+m)].value}*tau)")

	print('fifth loop')
	for n in range(0,3):
		#B and C terms in radians
		trigArgument = round(dataSht['D'+str(49+n)].value + dataSht['E'+str(49+n)].value*tau,7)
		lTerms[4] = round(lTerms[4] + dataSht['C'+str(49+n)].value*math.cos(trigArgument),0)
		print(f"{dataSht['C'+str(49+n)].value}*cos({dataSht['D'+str(49+n)].value}+{dataSht['E'+str(49+n)].value}*tau)")

	trigL5Arg = round(dataSht['D52'].value + dataSht['E52'].value*tau,6)
	lTerms[5] = round(dataSht['C52'].value*math.cos(trigL5Arg),6)

	print(lTerms)

	LTotal = 0.0
	for power in range(0,len(lTerms)):
		LTotal = LTotal + lTerms[power]*math.pow(tau,power)

	L = LTotal/math.pow(10,8)
	print(reduceAngle(math.degrees(L)))

	for a in range(0,9):
		#B and C terms are already in radians so it is not necessary to convert the entire argument to radians
		trigAgrument = dataSht['I'+str(2+a)].value + dataSht['J'+str(2+a)].value*tau
		bTerms[0] = bTerms[0] + dataSht['H'+str(2+a)].value*math.cos(trigAgrument)

	for b in range(0,4):
		#B and C terms are already in radians so it is not necessary to convert the entire argument to radians
		trigAgrument = dataSht['I'+str(11+b)].value + dataSht['J'+str(11+b)].value*tau
		bTerms[1] = bTerms[1] + dataSht['H'+str(11+b)].value*math.cos(trigAgrument)

	for c in range(0,4):
		#B and C terms are already in radians so it is not necessary to convert the entire argument to radians
		trigAgrument = dataSht['I'+str(15+c)].value + dataSht['J'+str(15+c)].value*tau
		bTerms[2] = bTerms[2] + dataSht['H'+str(15+c)].value*math.cos(trigAgrument)

	for d in range(0,4):
		#B and C terms are already in radians so it is not necessary to convert the entire argument to radians
		trigAgrument = dataSht['I'+str(19+d)].value + dataSht['J'+str(19+d)].value*tau
		bTerms[3] = bTerms[3] + dataSht['H'+str(19+d)].value*math.cos(trigAgrument)

	BTotal = 0.0
	for power in range(0,len(bTerms)):
		BTotal = BTotal + bTerms[power]*math.pow(tau,power)

	B = BTotal/math.pow(10,8)
	print(math.degrees(B))

	return [round(reduceAngle(math.degrees(L)),6),round(reduceAngle(math.degrees(B)),6)]

def venusRadiusVector(jd):
	dataWbk = load_workbook('/home/sdl5384/Desktop/Python_SRC/Astronomy/planetaryData.xlsx')
	dataSht = dataWbk.worksheets[1]
	tau = determine_tau(jd)

	rTerms = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
	for j in range(0,12):
		trigArgumentR0 = dataSht['N'+str(2+j)].value+dataSht['O'+str(2+j)].value*tau
		rTerms[0] = rTerms[0] + dataSht['M'+str(2+j)].value*math.cos(trigArgumentR0)

	for k in range(0,3):
		trigArgumentR1 = dataSht['N'+str(14+k)].value + dataSht['O'+str(14+k)].value*tau
		rTerms[1] = rTerms[1] + dataSht['M'+str(14+k)].value*math.cos(trigArgumentR1)

	for m in range(0,1):
		trigArgumentR2 = dataSht['N'+str(20+m)].value+dataSht['O'+str(20+m)].value*tau
		rTerms[2] = rTerms[2] + dataSht['M'+str(20+m)].value*math.cos(trigArgumentR2)

	for n in range(0,1):
		trigArgumentR3 = dataSht['N'+str(21+n)].value + dataSht['O'+str(21+n)].value*tau
		rTerms[3] = rTerms[3] + dataSht['M'+str(21+n)].value*math.cos(trigArgumentR3)

	rSum = 0.0
	for a in range(0,len(rTerms)):
		rSum = rSum + rTerms[a]*math.pow(tau,a)

	radius_vector = rSum/math.pow(10,8)

	return round(radius_vector,6)

def main():
	j = julianDay(12,20,1992)
	venus_coordinates = venus_heliocentric_coordinates(j)
	venus_radius = venusRadiusVector(j)
	#print(venus_coordinates)
	#print(venus_radius)

main()