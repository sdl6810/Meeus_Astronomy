import math
from astro_functions import reduceAngle, julianDay
from openpyxl import Workbook, load_workbook
from general_util import importXlColumnToPythonList

def determine_tau(julianDate):
	return round((julianDate - 2451545.0)/365250,6)

#all index limits in each for loop has been INCREASED BY ONE because Python does not include the last value in the process
#8/16/2023
def earth_heliocentric_coordinates(jd):
	dataWbk = load_workbook('/home/sdl5384/Desktop/Python_SRC/Astronomy/earthMoonData.xlsx')
	dataSht = dataWbk.worksheets[0]
	tau = determine_tau(jd)

	lTerms = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0] 

	bZeroSumTerms = 0.0
	bOneSumTerms = 0.0

	for i in range(0,64):
		#B and C terms are already in radians so it is not necessary to convert the entire argument to radians
		trigAgrument = dataSht['D'+str(2+i)].value + dataSht['E'+str(2+i)].value*tau
		lTerms[0] = lTerms[0] + dataSht['C'+str(2+i)].value*math.cos(trigAgrument)

	for j in range(0,34):
		#B and C terms already in radians.  No need to convert to radians
		trigAgrument = dataSht['D'+str(67+j)].value + dataSht['E'+str(67+j)].value*tau
		lTerms[1] = lTerms[1] + dataSht['C'+str(67+j)].value*math.cos(trigAgrument)

	for k in range(0,20):
		#B and C terms are already in radians.
		trigAgrument = dataSht['D'+str(102+k)].value + dataSht['E'+str(102+k)].value*tau
		lTerms[2] = lTerms[2] + dataSht['C'+str(102+k)].value*math.cos(trigAgrument)

	for m in range(0,7):
		#B and C terms are already in radians
		trigAgrument = dataSht['D'+str(123+m)].value + dataSht['E'+str(123+m)].value*tau
		lTerms[3] = lTerms[3] + dataSht['C'+str(123+m)].value*math.cos(trigAgrument)

	for n in range(0,3):
		#B and C terms in radians
		trigAgrument = dataSht['D'+str(131+n)].value + dataSht['E'+str(131+n)].value*tau
		lTerms[4] = lTerms[4] + dataSht['C'+str(131+n)].value*math.cos(trigAgrument)

	trigL5Arg = dataSht['D135'].value + dataSht['E135'].value*tau
	lTerms[5] = dataSht['C135'].value*math.cos(trigL5Arg)

	LTotal = 0.0
	for power in range(0,len(lTerms)):
		LTotal = LTotal + lTerms[power]*math.pow(tau,power)

	L = LTotal/math.pow(10,8)

	for a in range(0,5):
		#B and C are in radians
		trigArgumentB0 = dataSht['I'+str(2+a)].value + dataSht['J'+str(2+a)].value*tau
		bZeroSumTerms = bZeroSumTerms + dataSht['H'+str(2+a)].value*math.cos(trigArgumentB0)

	for b in range(0,2):
		trigArgumentB1 = dataSht['I'+str(8+b)].value + dataSht['J'+str(8+b)].value

	B = (bZeroSumTerms + bOneSumTerms)/math.pow(10,8)
	print(B)

	return [round(reduceAngle(math.degrees(L)),6),round(reduceAngle(math.degrees(B)),6)]

def radiusVector(jd):
	dataWbk = load_workbook('/home/sdl5384/Desktop/Python_SRC/Astronomy/earthMoonData.xlsx')
	dataSht = dataWbk.worksheets[0]
	tau = determine_tau(jd)

	rTerms = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
	for j in range(0,40):
		trigArgumentR0 = dataSht['N'+str(2+j)].value+dataSht['O'+str(2+j)].value*tau
		rTerms[0] = rTerms[0] + dataSht['M'+str(2+j)].value*math.cos(trigArgumentR0)


	for k in range(0,10):
		trigArgumentR1 = dataSht['N'+str(43+k)].value + dataSht['O'+str(43+k)].value*tau
		rTerms[1] = rTerms[1] + dataSht['M'+str(43+k)].value*math.cos(trigArgumentR1)

	for m in range(0,6):
		trigArgumentR2 = dataSht['N'+str(54+m)].value+dataSht['O'+str(54+m)].value*tau
		rTerms[2] = rTerms[2] + dataSht['M'+str(54+m)].value*math.cos(trigArgumentR2)

	for n in range(0,2):
		trigArgumentR3 = dataSht['N'+str(61+n)].value + dataSht['O'+str(61+n)].value*tau
		rTerms[3] = rTerms[3] + dataSht['M'+str(61+n)].value*math.cos(trigArgumentR3)

	trigArgumentR4 = dataSht['N64'].value + dataSht['O64'].value*tau
	rTerms[4] = dataSht['M64'].value*math.cos(trigArgumentR4)

	rSum = 0.0
	for a in range(0,len(rTerms)):
		rSum = rSum + rTerms[a]*math.pow(tau,a)

	radius_vector = rSum/math.pow(10,8)

	return round(radius_vector,6)

def main():
	jd = julianDay(8,13.25,2023)
	coordinates = earth_heliocentric_coordinates(jd)
	print(coordinates)

	r = radiusVector(jd)
	print(r)

main()